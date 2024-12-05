from pyecharts.charts import Line
from pyecharts import options as opts
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
import openpyxl


# 1. 加载 Excel 数据
excel_path = "AAAA.xlsx"  # 替换为您的文件路径
sheet_name = "成绩单"  # 替换为您的 Sheet 名称
wb = load_workbook(excel_path)
ws = wb[sheet_name]

# 获取数据
data = []
for row in ws.iter_rows(values_only=True):
    data.append(row)

# 转换为 DataFrame
header = data[0]  # 第一行为标题
content = data[1:]  # 其余行为内容
df = pd.DataFrame(content, columns=header)

# 2. 配置生成图片的文件夹
output_folder = "趋势图"
os.makedirs(output_folder, exist_ok=True)

# 3. 循环生成折线图
x_labels = df.columns[1:]  # 横坐标标签（第一单元到第五单元）
today = datetime.now().strftime("%Y%m%d")

for index, row in df.iterrows():
    name = row["姓名"]  # 获取姓名
    scores = row[1:].tolist()  # 获取分数

    # 最大值和最小值
    max_score = max(scores)
    min_score = min(scores)
    max_idx = scores.index(max_score)
    min_idx = scores.index(min_score)

    # 绘制折线图
    line = (
        Line()
        .add_xaxis(x_labels.tolist())
        .add_yaxis(
            series_name="分数",
            y_axis=scores,
            label_opts=opts.LabelOpts(is_show=True),  # 显示数据标签
        )
        .set_global_opts(
            title_opts=opts.TitleOpts(title=f"{name} 成绩趋势"),
            xaxis_opts=opts.AxisOpts(name="单元"),
            yaxis_opts=opts.AxisOpts(name="分数", max_=100, min_=0),
            tooltip_opts=opts.TooltipOpts(is_show=True),
        )
    )

    # 添加最大值和最小值的标注
    line.set_series_opts(
        markpoint_opts=opts.MarkPointOpts(
            data=[
                opts.MarkPointItem(name="最大值", coord=[x_labels[max_idx], max_score], value=max_score),
                opts.MarkPointItem(name="最小值", coord=[x_labels[min_idx], min_score], value=min_score),
            ]
        ),
        markline_opts=opts.MarkLineOpts(
            data=[
                opts.MarkLineItem(type_="max", name="最大值"),
                opts.MarkLineItem(type_="min", name="最小值"),
            ]
        ),
    )

    # 保存为 HTML 文件
    output_file = os.path.join(output_folder, f"{name}_成绩单_趋势_{today}.html")
    line.render(output_file)

print(f"折线图已生成并保存在文件夹: {output_folder}")