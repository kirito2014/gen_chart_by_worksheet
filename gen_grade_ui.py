import os
from tkinter import Tk, filedialog, StringVar, ttk, messagebox
from tkinter import Radiobutton
from ttkthemes import ThemedTk
from openpyxl import load_workbook
from pyecharts.charts import Line
from pyecharts import options as opts
import pandas as pd
from datetime import datetime
import matplotlib.pyplot as plt

# 配置matplotlib使用微软雅黑字体
import matplotlib
matplotlib.rcParams['font.family'] = 'Microsoft YaHei'

# 主程序逻辑
class TrendChartApp:
    def __init__(self, root):
        self.root = root
        self.root.title("成绩单趋势生成工具")
        self.root.geometry("600x450")
        self.root.configure(bg='#f0f0f0')  # 设置背景颜色
        self.root.set_theme("arc")  # 设置主题为arc
        self.root.option_add("*Font", "黑体 10")  # 设置全局字体

        # 初始化变量
        self.file_path = StringVar()
        self.sheet_name = StringVar()
        self.sheets = []
        self.file_format = StringVar(value="html")  # 默认选择html格式

        # 文件选择框
        ttk.Label(root, text="选择要生成的工作簿:").pack(pady=10)
        file_frame = ttk.Frame(root)
        file_frame.pack(pady=5)
        ttk.Entry(file_frame, textvariable=self.file_path, width=40).pack(side="left", padx=5)
        ttk.Button(file_frame, text="选择文件", command=self.select_file).pack(side="left")

        # Sheet选择框
        ttk.Label(root, text="选择你要生成的工作表:").pack(pady=10)
        self.sheet_dropdown = ttk.Combobox(root, textvariable=self.sheet_name, state="readonly", width=30)
        self.sheet_dropdown.pack(pady=5)

        # 选择文件格式
        ttk.Label(root, text="选择文件格式:").pack(pady=10)
        format_frame = ttk.Frame(root)
        format_frame.pack(pady=5)
        Radiobutton(format_frame, text="HTML", variable=self.file_format, value="html").pack(side="left")
        Radiobutton(format_frame, text="PNG", variable=self.file_format, value="png").pack(side="left")

        # 进度条
        self.progress = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
        self.progress.pack(pady=15)

        # 执行按钮和清除按钮在同一行
        button_frame = ttk.Frame(root)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="执行生成", command=self.generate_charts ).pack(side="left", padx=10)
        ttk.Button(button_frame, text="清除", command=self.clear ).pack(side="left", padx=10)

    def clear(self):
        self.file_path.set("")
        self.sheet_name.set("")
        self.sheets = []
        self.sheet_dropdown["values"] = self.sheets
        self.progress["value"] = 0
        self.root.update()

    def select_file(self):
        # 选择Excel文件
        file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
        if file_path:
            self.file_path.set(file_path)
            self.load_sheets()

    def load_sheets(self):
        # 加载文件中的Sheet
        try:
            wb = load_workbook(self.file_path.get())
            self.sheets = wb.sheetnames
            self.sheet_dropdown["values"] = self.sheets
            if self.sheets:
                self.sheet_name.set(self.sheets[0])
        except Exception as e:
            messagebox.showerror("错误", f"无法加载文件: {e}")

    def generate_charts(self):
        # 检查输入
        if not self.file_path.get():
            messagebox.showwarning("警告", "请先选择文件！")
            return
        if not self.sheet_name.get():
            messagebox.showwarning("警告", "请先选择Sheet！")
            return
        
        # 执行折线图生成
        try:
            self.progress["value"] = 0
            self.root.update()
            self.run_script()
        except Exception as e:
            messagebox.showerror("错误", f"生成失败: {e}")

    def run_script(self):
        # 生成折线图的脚本
        excel_path = self.file_path.get()
        sheet_name = self.sheet_name.get()
        file_format = self.file_format.get()

        # 加载Excel
        wb = load_workbook(excel_path)
        ws = wb[sheet_name]

        # 获取数据
        data = [row for row in ws.iter_rows(values_only=True)]
        header = data[0]  # 第一行标题
        content = data[1:]  # 其余行内容

        # 判断是否有序号列
        if header[0] == "序号":
            df = pd.DataFrame([row[1:] for row in content], columns=header[1:])  # 如果有序号，从第二列开始
        else:
            df = pd.DataFrame(content, columns=header)  # 否则照常处理

        # 配置保存目录
        output_folder = "趋势图"
        sub_folder = sheet_name
        output_folder_final = output_folder + "/" + sub_folder
        os.makedirs(output_folder, exist_ok=True)
        os.makedirs(output_folder_final, exist_ok=True)
        today = datetime.now().strftime("%Y%m%d")
        total_rows = len(df)
        self.progress["maximum"] = total_rows

        x_labels = df.columns[1:]  # 横坐标标签
        for index, row in df.iterrows():
            name = row["姓名"]
            scores = row[1:].tolist()

            # 最大值与最小值
            max_score = max(scores)
            min_score = min(scores)
            max_idx = scores.index(max_score)
            min_idx = scores.index(min_score)

            # 使用pyecharts生成折线图
            line = (
                Line()
                .add_xaxis(x_labels.tolist())
                .add_yaxis("分数", scores, is_smooth=True, label_opts=opts.LabelOpts(is_show=True))
                .set_global_opts(
                    title_opts=opts.TitleOpts(title=f"{name} 成绩趋势"),
                    xaxis_opts=opts.AxisOpts(name="单元"),
                    yaxis_opts=opts.AxisOpts(name="分数", max_=100, min_=0),
                    tooltip_opts=opts.TooltipOpts(is_show=True),
                )
                .set_series_opts(
                    markpoint_opts=opts.MarkPointOpts(
                        data=[ 
                            opts.MarkPointItem(name="最大值", coord=[x_labels[max_idx], max_score], value=max_score),
                            opts.MarkPointItem(name="最小值", coord=[x_labels[min_idx], min_score], value=min_score),
                        ]
                    ),
                    markline_opts=opts.MarkLineOpts(
                        data=[
                            opts.MarkLineItem(type_="max", name="最大值"),
                            opts.MarkLineItem(type_="min", name="最小值"),
                        ]
                    ),
                )
            )

            if file_format == "html":
                # 保存为html
                output_file_html = os.path.join(output_folder_final, f"{name}_成绩单_趋势_{today}.html")
                line.render(output_file_html)
            else:
                # 使用matplotlib生成PNG图像
                plt.figure(figsize=(10, 6))
                plt.plot(x_labels.tolist(), scores, marker='o', linestyle='-', color='b', label="分数")
                plt.title(f"{name} 成绩趋势")
                plt.xlabel("单元")
                #平滑曲线
                plt.xticks(rotation=45)
                plt.ylabel("分数")
                plt.ylim(0, 100)
                plt.grid(True)
                plt.legend()
                
                # 保存为PNG
                output_file_png = os.path.join(output_folder_final, f"{name}_成绩单_趋势_{today}.png")
                plt.savefig(output_file_png, format='png')
                plt.close()  # 关闭图形

            # 更新进度条
            self.progress["value"] += 1
            self.root.update()

        messagebox.showinfo("完成", f"所有图表已生成！图表保存在文件夹: {output_folder_final}")

if __name__ == "__main__":
    root = ThemedTk(theme="arc")  # 使用ttkthemes美化
    app = TrendChartApp(root)
    root.mainloop()
